from datetime import datetime, timedelta
from jose import JWTError, jwt
from dotenv import load_dotenv
from passlib.context import CryptContext
from typing import List, Tuple
import csv
import json
import io

import os

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

load_dotenv()

SECRET_KEY = os.getenv("SECRET_KEY")
ALGORITHM = os.getenv("ALGORITHM")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 60))

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def hash_password(password: str) -> str:
    """Hashes the user password safely."""
    if len(password.encode("utf-8")) > 72:
        password = password[:72]
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verifies a password against its hash."""
    return pwd_context.verify(plain_password[:72], hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(token: str):
    print("🔹 Raw token received:", token)   # 👈 Add this
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        print("✅ DECODED PAYLOAD:", payload)
        email: str = payload.get("sub")
        if email is None:
            print("⚠️ No 'sub' in payload")
            return None
        return email
    except JWTError as e:
        print("❌ JWT ERROR:", e)
        return None

def get_current_admin_user(token: str, db):
    """Verify token and return admin user if authenticated and is_admin=True"""
    from . import models
    email = verify_token(token)
    if email is None:
        return None
    user = db.query(models.User).filter(models.User.email == email).first()
    if user and user.is_admin:
        return user
    return None


# Upload Utilities
def parse_csv_file(file) -> List[dict]:
    """Parse CSV file and return list of dictionaries"""
    content = file.file.read()
    file.file.seek(0)  # Reset file pointer
    text = content.decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(text))
    return list(csv_reader)


def parse_json_file(file) -> List[dict]:
    """Parse JSON file and return list of dictionaries"""
    content = file.file.read()
    file.file.seek(0)  # Reset file pointer
    text = content.decode('utf-8')
    data = json.loads(text)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        return [data]
    else:
        raise ValueError("JSON file must contain an array or object")


def parse_excel_file(file) -> List[dict]:
    """Parse Excel file (.xlsx) and return list of dictionaries"""
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl is not installed. Please install it to support Excel files.")
    
    content = file.file.read()
    file.file.seek(0)  # Reset file pointer
    
    # Load workbook from bytes
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    
    # Get the first sheet
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        header_value = cell.value
        if header_value is None:
            break
        # Normalize header: strip whitespace, convert to lowercase, replace spaces with underscores
        normalized = str(header_value).strip().lower().replace(' ', '_').replace('-', '_')
        headers.append(normalized)
    
    if not headers:
        raise ValueError("Excel file must have headers in the first row")
    
    # Read data rows
    data_list = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        # Check if row is empty
        if all(cell.value is None or str(cell.value).strip() == '' for cell in row[:len(headers)]):
            continue
        
        row_data = {}
        for idx, cell in enumerate(row[:len(headers)]):
            header = headers[idx]
            value = cell.value
            
            # Convert None or empty strings to None
            if value is None or (isinstance(value, str) and value.strip() == ''):
                row_data[header] = None
            else:
                # Try to preserve the original value type
                row_data[header] = value
        
        # Only add non-empty rows
        if any(v is not None for v in row_data.values()):
            data_list.append(row_data)
    
    return data_list


def validate_policy_data(data: dict, db) -> Tuple[bool, str]:
    """Validate policy data (check provider_id, type_id exist)"""
    from . import models
    if 'provider_id' not in data:
        return False, "Missing required field: provider_id"
    if 'type_id' not in data:
        return False, "Missing required field: type_id"
    if 'name' not in data:
        return False, "Missing required field: name"
    
    provider = db.query(models.Provider).filter(models.Provider.provider_id == int(data['provider_id'])).first()
    if not provider:
        return False, f"Provider with ID {data['provider_id']} not found"
    
    insurance_type = db.query(models.InsuranceType).filter(models.InsuranceType.type_id == int(data['type_id'])).first()
    if not insurance_type:
        return False, f"Insurance type with ID {data['type_id']} not found"
    
    return True, ""


def validate_tariff_data(data: dict, db) -> Tuple[bool, str]:
    """Validate tariff data (check policy_id exists)"""
    from . import models
    if 'policy_id' not in data or data['policy_id'] is None:
        return False, "Missing required field: policy_id"
    if 'age_min' not in data or data['age_min'] is None:
        return False, "Missing required field: age_min"
    if 'age_max' not in data or data['age_max'] is None:
        return False, "Missing required field: age_max"
    if 'class_type' not in data or not data['class_type']:
        return False, "Missing required field: class_type"
    # family_min and family_max are required but have defaults, so we check if they exist or use defaults
    if 'family_min' not in data or data['family_min'] is None:
        data['family_min'] = 1
    if 'family_max' not in data or data['family_max'] is None:
        data['family_max'] = 1
    
    # Convert policy_id to int if it's not already
    try:
        policy_id = int(data['policy_id']) if not isinstance(data['policy_id'], int) else data['policy_id']
    except (ValueError, TypeError):
        return False, f"Invalid policy_id: {data['policy_id']}"
    
    policy = db.query(models.InsurancePlan).filter(models.InsurancePlan.policy_id == policy_id).first()
    if not policy:
        return False, f"Policy with ID {policy_id} not found"
    
    return True, ""


def validate_criteria_data(data: dict) -> Tuple[bool, str]:
    """Validate criteria_data structure (in-patient) against schema"""
    from . import schemas
    try:
        # Try to parse as InPatientCriteriaData
        schemas.InPatientCriteriaData(**data)
        return True, ""
    except Exception as e:
        return False, f"Invalid criteria_data structure: {str(e)}"


def validate_outpatient_criteria_data(data: dict) -> Tuple[bool, str]:
    """Validate outpatient_criteria_data structure against schema"""
    from . import schemas
    try:
        # Try to parse as OutPatientCriteriaData
        schemas.OutPatientCriteriaData(**data)
        return True, ""
    except Exception as e:
        return False, f"Invalid outpatient_criteria_data structure: {str(e)}"


def parse_criteria_excel_to_json(data_list: List[dict]) -> List[dict]:
    """Convert flat Excel data structure to nested JSON structure for criteria"""
    # Define all coverage fields
    IN_PATIENT_GENERAL = [
        "annual_limit", "scope_of_coverage", "network", "geographic_coverage_elective",
        "geographic_coverage_emergency", "waiting_period", "non_direct_billing", "cold_case",
        "hospital_accommodation", "road_ambulance", "maternity_in_patient", "maternity_lab_test",
        "new_born", "nursery_incubator", "extra_bed_parent", "home_care",
        "plan_upgrade_downgrade", "passive_war", "payment_frequency", "pre_existing_conditions",
    ]
    
    IN_PATIENT_CASE = [
        "physiotherapy", "work_related_injuries", "acute_allergy_treatments", "bariatric_surgeries",
        "breast_reconstruction", "chemotherapy_radiotherapy", "chronic_conditions",
        "congenital_cases_lifetime", "congenital_tests_thalassemia", "epidural", "epilepsy", "icu",
        "infertility_impotence_sterility", "laparoscopic_procedures", "migraines", "motorcycling",
        "organ_transplant", "polysomnography", "prosthesis_due_to_accident",
        "prosthesis_due_to_sickness", "rehabilitation", "renal_dialysis", "scoliosis",
        "std_excluding_hiv", "varicocele", "varicose_veins", "morgue_burial_expenses",
        "genetic_tests", "diagnostic_tests", "ambulatory_laboratory_exams",
        "doctor_visits_consultations", "prescribed_medicines_drugs",
    ]
    
    OUT_PATIENT = [
        "outpatient_annual_limit", "outpatient_coverage", "outpatient_network",
        "outpatient_deductible", "diagnostic_tests", "ambulatory_laboratory_exams",
        "doctor_visits_consultations", "prescribed_medicines_drugs",
    ]
    
    result = []
    
    for row_data in data_list:
        if 'policy_id' not in row_data or row_data['policy_id'] is None:
            continue
        
        # Build in-patient general coverages
        general_coverages = {}
        for field in IN_PATIENT_GENERAL:
            # Try to find the column - template uses "In-Patient General: {Field} - Notes"
            notes_key = None
            field_name_normalized = field.replace('_', ' ').title()
            
            for key in row_data.keys():
                normalized_key = key.lower()
                # Match patterns like "in-patient general: annual limit - notes" or variations
                if ("in" in normalized_key and "patient" in normalized_key and "general" in normalized_key) or \
                   ("general" in normalized_key):
                    # Extract the field name from the column header
                    key_normalized = key.replace('-', ' ').replace('_', ' ').lower()
                    field_normalized = field.replace('_', ' ').lower()
                    # Check if the field name is in the column header
                    if field_normalized in key_normalized:
                        notes_key = key
                        break
            
            notes_value = row_data.get(notes_key, "") if notes_key else ""
            if notes_value is None:
                notes_value = ""
            general_coverages[field] = {"notes": str(notes_value).strip()}
        
        # Build in-patient case coverages
        case_coverages = {}
        for field in IN_PATIENT_CASE:
            notes_key = None
            field_name_normalized = field.replace('_', ' ').title()
            
            for key in row_data.keys():
                normalized_key = key.lower()
                if ("in" in normalized_key and "patient" in normalized_key and "case" in normalized_key) or \
                   ("case" in normalized_key and "general" not in normalized_key):
                    key_normalized = key.replace('-', ' ').replace('_', ' ').lower()
                    field_normalized = field.replace('_', ' ').lower()
                    if field_normalized in key_normalized:
                        notes_key = key
                        break
            
            notes_value = row_data.get(notes_key, "") if notes_key else ""
            if notes_value is None:
                notes_value = ""
            case_coverages[field] = {"notes": str(notes_value).strip()}
        
        # Build out-patient coverages
        out_patient_coverages = {}
        for field in OUT_PATIENT:
            notes_key = None
            
            for key in row_data.keys():
                normalized_key = key.lower()
                if ("out" in normalized_key and "patient" in normalized_key) or \
                   ("outpatient" in normalized_key):
                    key_normalized = key.replace('-', ' ').replace('_', ' ').lower()
                    field_normalized = field.replace('outpatient_', '').replace('_', ' ').lower()
                    # Handle special case where field already contains "outpatient"
                    if field.startswith("outpatient_"):
                        field_normalized = field.replace('_', ' ').lower()
                    if field_normalized in key_normalized:
                        notes_key = key
                        break
            
            notes_value = row_data.get(notes_key, "") if notes_key else ""
            if notes_value is None:
                notes_value = ""
            out_patient_coverages[field] = {"notes": str(notes_value).strip()}
        
        # Build the nested structure
        criteria_data = {
            "in_patient": {
                "general_coverages": general_coverages,
                "case_coverages": case_coverages
            }
        }
        
        outpatient_criteria_data = {
            "out_patient": out_patient_coverages
        }
        
        result.append({
            "policy_id": int(row_data['policy_id']),
            "criteria_data": criteria_data,
            "outpatient_criteria_data": outpatient_criteria_data
        })
    
    return result