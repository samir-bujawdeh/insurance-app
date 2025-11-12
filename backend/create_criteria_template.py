"""
Script to create an Excel template for plan criteria uploads
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Define all coverage fields
IN_PATIENT_GENERAL = [
    "annual_limit",
    "scope_of_coverage",
    "network",
    "geographic_coverage_elective",
    "geographic_coverage_emergency",
    "waiting_period",
    "non_direct_billing",
    "cold_case",
    "hospital_accommodation",
    "road_ambulance",
    "maternity_in_patient",
    "maternity_lab_test",
    "new_born",
    "nursery_incubator",
    "extra_bed_parent",
    "home_care",
    "plan_upgrade_downgrade",
    "passive_war",
    "payment_frequency",
    "pre_existing_conditions",
]

IN_PATIENT_CASE = [
    "physiotherapy",
    "work_related_injuries",
    "acute_allergy_treatments",
    "bariatric_surgeries",
    "breast_reconstruction",
    "chemotherapy_radiotherapy",
    "chronic_conditions",
    "congenital_cases_lifetime",
    "congenital_tests_thalassemia",
    "epidural",
    "epilepsy",
    "icu",
    "infertility_impotence_sterility",
    "laparoscopic_procedures",
    "migraines",
    "motorcycling",
    "organ_transplant",
    "polysomnography",
    "prosthesis_due_to_accident",
    "prosthesis_due_to_sickness",
    "rehabilitation",
    "renal_dialysis",
    "scoliosis",
    "std_excluding_hiv",
    "varicocele",
    "varicose_veins",
    "morgue_burial_expenses",
    "genetic_tests",
    "diagnostic_tests",
    "ambulatory_laboratory_exams",
    "doctor_visits_consultations",
    "prescribed_medicines_drugs",
]

OUT_PATIENT = [
    "outpatient_annual_limit",
    "outpatient_coverage",
    "outpatient_network",
    "outpatient_deductible",
    "diagnostic_tests",
    "ambulatory_laboratory_exams",
    "doctor_visits_consultations",
    "prescribed_medicines_drugs",
]

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan Criteria"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Create headers
    headers = ["Policy ID"]
    
    # Add in-patient general coverage headers
    for field in IN_PATIENT_GENERAL:
        headers.append(f"In-Patient General: {field.replace('_', ' ').title()} - Notes")
    
    # Add in-patient case coverage headers
    for field in IN_PATIENT_CASE:
        headers.append(f"In-Patient Case: {field.replace('_', ' ').title()} - Notes")
    
    # Add out-patient coverage headers
    for field in OUT_PATIENT:
        headers.append(f"Out-Patient: {field.replace('_', ' ').title()} - Notes")
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Policy ID
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 35
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Add example row
    example_row = [1] + [""] * (len(headers) - 1)
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        if col_idx == 1:
            cell.font = Font(italic=True, color="808080")
            cell.value = "Example: 1"
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 0)
    instructions = [
        ["PLAN CRITERIA UPLOAD TEMPLATE - INSTRUCTIONS"],
        [""],
        ["1. POLICY ID"],
        ["   - Enter the numeric Policy ID for each plan"],
        ["   - This must match an existing policy in the system"],
        [""],
        ["2. NOTES FIELDS"],
        ["   - All coverage items only require notes"],
        ["   - Enter any relevant notes or descriptions for each coverage item"],
        ["   - Leave blank if no notes are needed"],
        [""],
        ["3. IN-PATIENT GENERAL COVERAGES"],
        ["   - These are general coverage items for in-patient services"],
        ["   - 20 different coverage types"],
        [""],
        ["4. IN-PATIENT CASE COVERAGES"],
        ["   - These are specific case-based coverage items"],
        ["   - 32 different coverage types"],
        [""],
        ["5. OUT-PATIENT COVERAGES"],
        ["   - These are coverage items for out-patient services"],
        ["   - 8 different coverage types"],
        [""],
        ["6. UPLOAD"],
        ["   - Save this file as .xlsx format"],
        ["   - Upload through the admin dashboard"],
        ["   - Each row represents criteria for one policy"],
        [""],
        ["TOTAL COLUMNS: " + str(len(headers))],
        ["   - 1 Policy ID column"],
        ["   - " + str(len(IN_PATIENT_GENERAL)) + " In-Patient General coverage columns"],
        ["   - " + str(len(IN_PATIENT_CASE)) + " In-Patient Case coverage columns"],
        ["   - " + str(len(OUT_PATIENT)) + " Out-Patient coverage columns"],
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction[0] and instruction[0][0].isdigit():
            cell.font = Font(bold=True)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Save the file
    filename = "plan_criteria_template.xlsx"
    wb.save(filename)
    print(f"Template created: {filename}")
    print(f"Total columns: {len(headers)}")
    print(f"  - Policy ID: 1")
    print(f"  - In-Patient General: {len(IN_PATIENT_GENERAL)}")
    print(f"  - In-Patient Case: {len(IN_PATIENT_CASE)}")
    print(f"  - Out-Patient: {len(OUT_PATIENT)}")

if __name__ == "__main__":
    create_template()

